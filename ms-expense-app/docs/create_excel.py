#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""マネーフォワード vs freee 比較資料 Excel生成スクリプト（達人連携前提版）"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ワークブック作成
wb = Workbook()

# スタイル定義
header_fill_mf = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
header_fill_freee = PatternFill(start_color="E6F7ED", end_color="E6F7ED", fill_type="solid")
header_fill_gray = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
header_fill_yellow = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

font_header = Font(bold=True, size=11)
font_title = Font(bold=True, size=14)
font_mf = Font(bold=True, color="3B7DE9")
font_freee = Font(bold=True, color="00AA4F")

# ========== シート1: 総合比較 ==========
ws1 = wb.active
ws1.title = "総合比較"

ws1.merge_cells('A1:D1')
ws1['A1'] = "マネーフォワード vs freee 総合比較表"
ws1['A1'].font = Font(bold=True, size=16)
ws1['A1'].alignment = Alignment(horizontal='center')

ws1['A2'] = "※マネーフォワードは達人シリーズ連携を前提として評価"
ws1['A2'].font = Font(italic=True, color="666666")

headers = ["比較項目", "マネーフォワード\n（+達人シリーズ）", "freee\n（freee申告）", "備考"]
for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=4, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data = [
    ["記帳効率（大量取引）", "★★★☆☆", "★★★★★", "freeeは自動登録機能あり"],
    ["会計事務所との親和性", "★★★★★", "★★★☆☆", "MFは従来の複式簿記UI"],
    ["税務申告機能", "★★★★★", "★★★★★", "MF+達人、freee申告 共に充実"],
    ["内訳書・概況書自動化", "★★★★★", "★★★★☆", "MF+達人は完全自動連携可"],
    ["償却資産申告", "★★★★★", "★★★★★", "達人/freee申告共に対応"],
    ["給与・労務一体性", "★★★★☆", "★★★★★", "freeeはオールインワン"],
    ["初心者の使いやすさ", "★★★☆☆", "★★★★★", "freeeは簿記知識不要"],
    ["拡張性・API", "★★★☆☆", "★★★★★", "freeeはAPI完全公開"],
    ["総コスト（申告込み）", "★★★☆☆", "★★★★☆", "達人ライセンス費用考慮"],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if col_idx == 2:
            cell.font = font_mf
        elif col_idx == 3:
            cell.font = font_freee

ws1.column_dimensions['A'].width = 25
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 32
ws1.row_dimensions[4].height = 35

# ========== シート2: 会計入力・記帳 ==========
ws2 = wb.create_sheet("会計入力・記帳")

ws2.merge_cells('A1:D1')
ws2['A1'] = "会計入力・記帳効率の比較"
ws2['A1'].font = Font(bold=True, size=14)

headers = ["項目", "マネーフォワード", "freee", "評価"]
for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data = [
    ["UI設計思想", "従来の複式簿記に準拠\n経理経験者・税理士向け", "簿記知識不要の独自UI\n初心者・スタートアップ向け", "用途による"],
    ["仕訳入力方式", "貸借科目を明示する従来型\n仕訳帳イメージで入力", "取引登録型（家計簿感覚）\n貸借を意識させない設計", "用途による"],
    ["自動仕訳", "推測までで「登録ボタン」が必要\n一括登録は可能", "ルール設定で完全自動登録可能\n大量取引に強み", "freee優位"],
    ["銀行連携数", "◎ 2,300以上\n未確定明細も取込可能", "○ 主要金融機関対応", "MF優位"],
    ["AI学習機能", "○ 科目推測の精度向上", "○ 科目推測の精度向上", "同等"],
    ["クレジットカード連携", "◎ 未確定明細も取込可能\nリアルタイム性が高い", "○ 確定明細のみ", "MF優位"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws2.column_dimensions['A'].width = 20
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 35
ws2.column_dimensions['D'].width = 15

# ========== シート3: 税務申告（達人連携前提で再評価） ==========
ws3 = wb.create_sheet("税務申告")

ws3.merge_cells('A1:D1')
ws3['A1'] = "税務申告対応の比較（達人シリーズ連携前提）"
ws3['A1'].font = Font(bold=True, size=14)

ws3['A2'] = "※マネーフォワードは達人シリーズとの連携を前提として評価"
ws3['A2'].font = Font(italic=True, color="666666")

headers = ["項目", "マネーフォワード\n（+達人シリーズ）", "freee\n（freee申告）", "評価"]
for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=4, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data = [
    ["法人税申告", "◎ 法人税の達人\nMFからAPI連携でデータ取込\n電子申告対応", "◎ freee申告（法人税）\n年額29,800円〜\n電子申告対応", "同等"],
    ["消費税申告", "◎ 消費税の達人\nMFから消費税集計データ連携\nインボイス対応", "◎ freee申告（消費税）\nfreee会計に付帯\nインボイス対応", "同等"],
    ["勘定科目内訳明細書", "◎ 内訳概況書の達人\nMFから勘定科目・補助科目\n残高を自動連携", "○ freee申告 内訳書・概況書\n試算表から連携可能\n※別途契約必要", "MF+達人優位"],
    ["法人事業概況説明書", "◎ 内訳概況書の達人\nMFから月別残高等を自動連携", "○ freee申告 内訳書・概況書\n会計データ連携可能\n※別途契約必要", "MF+達人優位"],
    ["所得税確定申告\n（個人事業主）", "○ クラウド確定申告\ne-Tax対応", "◎ freee会計 確定申告機能\nスマホのみで申告可能", "freee優位"],
    ["所得税申告\n（税理士向け代理）", "◎ 所得税の達人\n代理申告対応", "◎ freee申告（所得税）\nアドバイザー向け\n一括電子申告可", "同等"],
    ["電子申告", "◎ 達人シリーズで一括対応\ne-Tax/eLTAX完全対応", "◎ freee申告内で完結\n法人税+消費税一括送信可", "同等"],
    ["申告書作成の効率", "◎ 達人シリーズの実績と安定性\n会計事務所での標準ツール", "○ freee内で完結\n操作習熟が必要", "MF+達人優位"],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws3['A14'] = "【結論】達人シリーズを導入済みの会計事務所であれば、MF+達人の組み合わせは申し分ない。"
ws3['A15'] = "内訳書・概況書の自動連携もMFから達人へ科目残高・月別残高を連携でき、効率的に作成可能。"
ws3['A16'] = "freee申告も機能は充実しているが、内訳書・概況書は別途契約が必要な点に注意。"

ws3.column_dimensions['A'].width = 24
ws3.column_dimensions['B'].width = 35
ws3.column_dimensions['C'].width = 35
ws3.column_dimensions['D'].width = 18
ws3.row_dimensions[4].height = 35

# ========== シート4: MF+達人連携詳細 ==========
ws3b = wb.create_sheet("MF+達人連携詳細")

ws3b.merge_cells('A1:C1')
ws3b['A1'] = "マネーフォワード → 達人シリーズ 連携機能一覧"
ws3b['A1'].font = Font(bold=True, size=14)

headers = ["連携先達人ソフト", "連携データ", "連携方法"]
for col, header in enumerate(headers, 1):
    cell = ws3b.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_mf
    cell.border = thin_border

data = [
    ["法人税の達人", "決算書データ\n（貸借対照表・損益計算書・株主資本等変動計算書）", "API連携 / 中間ファイル出力"],
    ["消費税の達人", "消費税集計データ\n（課税売上・課税仕入等）", "API連携 / 中間ファイル出力"],
    ["内訳概況書の達人\n（内訳書用）", "勘定科目名・補助科目名\n決算修正後の科目残高・補助科目残高\n取引先別残高", "API連携 / 中間ファイル出力"],
    ["内訳概況書の達人\n（概況書用）", "各科目の決算残高\n月別残高", "API連携 / 中間ファイル出力"],
    ["減価償却の達人", "固定資産台帳データ", "中間ファイル出力"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3b.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws3b['A10'] = "【API連携の設定】"
ws3b['A10'].font = Font(bold=True)
ws3b['A11'] = "・マネーフォワード クラウドのアプリポータルの利用開始が必要"
ws3b['A12'] = "・連携操作を行うユーザーへの権限付与が必要"
ws3b['A13'] = "・達人側から「データ取込」操作でMFのデータを自動取得"

ws3b['A15'] = "【内訳書・概況書の作成フロー】"
ws3b['A15'].font = Font(bold=True)
ws3b['A16'] = "1. MFで決算整理仕訳まで完了"
ws3b['A17'] = "2. MF画面から［決算・申告］→［達人シリーズ連携］を選択"
ws3b['A18'] = "3. 中間ファイルを出力 または API連携でデータ送信"
ws3b['A19'] = "4. 内訳概況書の達人で「残高取込」を実行"
ws3b['A20'] = "5. 科目残高・取引先別残高が自動で各帳票に反映"
ws3b['A21'] = "6. 必要に応じて手修正後、PDF出力または電子申告"

ws3b.column_dimensions['A'].width = 25
ws3b.column_dimensions['B'].width = 40
ws3b.column_dimensions['C'].width = 28

# ========== シート5: freee申告詳細 ==========
ws3c = wb.create_sheet("freee申告詳細")

ws3c.merge_cells('A1:C1')
ws3c['A1'] = "freee申告 プラン・機能一覧"
ws3c['A1'].font = Font(bold=True, size=14)

headers = ["プラン名", "年額（税別）", "主な機能"]
for col, header in enumerate(headers, 1):
    cell = ws3c.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_freee
    cell.border = thin_border

data = [
    ["freee申告 法人税スターター", "29,800円", "法人税確定申告\n電子申告対応"],
    ["freee申告 法人税スタンダード", "48,800円", "法人税確定申告\n+高度な機能"],
    ["freee申告 法人税アドバンス", "要問合せ", "資本金1億円超対応\n固定資産100件以上対応"],
    ["freee申告 消費税", "freee会計に付帯", "消費税申告書作成\n電子申告・インボイス対応"],
    ["freee申告 所得税", "アドバイザー向け", "税理士の代理申告用\n一括電子申告可"],
    ["freee申告 内訳書・概況書", "別途契約", "勘定科目内訳明細書\n法人事業概況説明書\n※電子申告機能なし"],
    ["freee申告 償却資産", "59,800円", "償却資産申告書作成\neLTAX電子申告対応"],
    ["freee申告 年調・法定調書", "freee人事労務に付帯", "年末調整\n法定調書合計表"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws3c.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws3c['A14'] = "【freee申告 内訳書・概況書の注意点】"
ws3c['A14'].font = Font(bold=True)
ws3c['A15'] = "・freee会計の試算表・在庫棚卸・固定資産台帳から連携可能"
ws3c['A16'] = "・連携実行時に編集情報が初期化されるため、連携後の手修正に注意"
ws3c['A17'] = "・取引先別残高連携には取引先マスタの設定が必要"
ws3c['A18'] = "・単体での電子申告機能なし（法人税申告書の添付書類として出力）"

ws3c.column_dimensions['A'].width = 30
ws3c.column_dimensions['B'].width = 22
ws3c.column_dimensions['C'].width = 35

# ========== シート6: 償却資産申告 ==========
ws4 = wb.create_sheet("償却資産申告")

ws4.merge_cells('A1:D1')
ws4['A1'] = "償却資産申告の比較"
ws4['A1'].font = Font(bold=True, size=14)

headers = ["項目", "マネーフォワード\n（+達人シリーズ）", "freee\n（freee申告）", "評価"]
for col, header in enumerate(headers, 1):
    cell = ws4.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

data = [
    ["固定資産台帳", "○ 標準機能で対応", "○ 標準機能で対応", "同等"],
    ["減価償却自動計上", "○ 毎期自動仕訳", "○ 毎期自動仕訳", "同等"],
    ["償却資産申告書作成", "◎ 減価償却の達人\nMFから固定資産データ連携", "◎ freee申告 償却資産\nソフト内で申告書作成可能", "同等"],
    ["電子申告（eLTAX）", "◎ 達人シリーズで対応", "◎ freee内で完結\neLTAXインストール不要", "同等"],
    ["追加費用", "達人シリーズライセンス費用", "年額59,800円（税別）\n※アドバイザー契約者は追加不要", "条件による"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws4.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws4.column_dimensions['A'].width = 22
ws4.column_dimensions['B'].width = 32
ws4.column_dimensions['C'].width = 32
ws4.column_dimensions['D'].width = 15
ws4.row_dimensions[3].height = 35

# ========== シート7: 給与・労務 ==========
ws5 = wb.create_sheet("給与・年末調整・労務")

ws5.merge_cells('A1:D1')
ws5['A1'] = "給与・年末調整・労務・社保・法定調書の比較"
ws5['A1'].font = Font(bold=True, size=14)

headers = ["項目", "マネーフォワード", "freee", "評価"]
for col, header in enumerate(headers, 1):
    cell = ws5.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data = [
    ["給与計算", "クラウド給与\n勤怠連携対応", "人事労務freee\n会計と一体型", "freee優位"],
    ["年末調整", "クラウド年末調整\nWeb完結・e-Tax対応", "freee申告 年調・法定調書\n人事労務に付帯", "同等"],
    ["社会保険手続き", "クラウド社会保険\n算定基礎届・月変届出力", "人事労務freee内\n算定基礎届・月変届出力", "同等"],
    ["法定調書", "○ 給与支払報告書・源泉徴収票\ne-Tax/eLTAX連携", "○ freee申告 年調・法定調書\n法定調書合計表まで対応", "同等"],
    ["外部連携", "◎ SmartHR、Touch On Time等", "○ 主要サービス対応", "MF優位"],
    ["料金", "基本料金内 or 追加契約", "月額2,600円〜（5名まで）", "条件による"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws5.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 35
ws5.column_dimensions['C'].width = 35
ws5.column_dimensions['D'].width = 15

# ========== シート8: 経費精算 ==========
ws6 = wb.create_sheet("経費精算")

ws6.merge_cells('A1:D1')
ws6['A1'] = "経費精算の比較"
ws6['A1'].font = Font(bold=True, size=14)

headers = ["項目", "マネーフォワード", "freee", "評価"]
for col, header in enumerate(headers, 1):
    cell = ws6.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data = [
    ["サービス形態", "クラウド経費（別契約）", "freee会計内蔵 / freee経費精算", "freee優位"],
    ["領収書読取", "◎ AI-OCR高精度\n使うほど学習", "◎ AI-OCR対応\n税区分・金額自動入力", "同等"],
    ["申請・承認", "○ スマホ完結", "◎ スマホ・LINE対応", "freee優位"],
    ["会計連携", "○ 仕訳自動計上", "◎ シームレス連携", "freee優位"],
    ["振込データ作成", "○ 対応", "○ 対応", "同等"],
    ["電子帳簿保存法", "○ JIIMA認証取得", "○ JIIMA認証取得", "同等"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws6.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 35
ws6.column_dimensions['C'].width = 35
ws6.column_dimensions['D'].width = 15

# ========== シート9: 請求・支払 ==========
ws7 = wb.create_sheet("請求・支払")

ws7.merge_cells('A1:D1')
ws7['A1'] = "請求・支払の比較"
ws7['A1'].font = Font(bold=True, size=14)

headers = ["項目", "マネーフォワード", "freee", "評価"]
for col, header in enumerate(headers, 1):
    cell = ws7.cell(row=3, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data = [
    ["請求書発行", "クラウド請求書（別契約）", "freee請求書（内蔵機能）", "freee優位"],
    ["見積書作成", "○ 対応", "○ 対応", "同等"],
    ["デザインテンプレート", "標準テンプレート", "◎ 40種類以上", "freee優位"],
    ["定期請求・一括請求", "○ 対応", "○ 上位プランで対応", "同等"],
    ["入金消込・督促", "◎ 自動化機能あり", "○ 対応", "MF優位"],
    ["会計連携", "◎ 仕訳自動連携", "◎ シームレス連携", "同等"],
    ["インボイス制度対応", "○ 適格請求書対応", "○ 適格請求書対応", "同等"],
]

for row_idx, row_data in enumerate(data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws7.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws7.column_dimensions['A'].width = 22
ws7.column_dimensions['B'].width = 35
ws7.column_dimensions['C'].width = 35
ws7.column_dimensions['D'].width = 15

# ========== シート10: 料金プラン ==========
ws8 = wb.create_sheet("料金プラン")

ws8.merge_cells('A1:E1')
ws8['A1'] = "料金プラン比較（法人向け・税抜）"
ws8['A1'].font = Font(bold=True, size=14)

# 会計ソフト料金
ws8['A3'] = "【会計ソフト】"
ws8['A3'].font = font_header

headers = ["プラン", "MF月額", "MF年額", "freee月額", "freee年額"]
for col, header in enumerate(headers, 1):
    cell = ws8.cell(row=4, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

data = [
    ["スモール/スターター", "3,980円", "35,760円", "4,780円", "47,760円"],
    ["ビジネス/スタンダード", "5,980円", "59,760円", "9,280円", "95,760円"],
    ["エンタープライズ等", "要問合せ", "要問合せ", "要問合せ", "要問合せ"],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws8.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

# 申告ソフト料金
ws8['A10'] = "【税務申告ソフト】"
ws8['A10'].font = font_header

headers2 = ["項目", "MF + 達人シリーズ", "freee申告"]
for col, header in enumerate(headers2, 1):
    cell = ws8.cell(row=11, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data2 = [
    ["法人税申告", "法人税の達人\n（ライセンス費用別途）", "freee申告 年額29,800円〜"],
    ["消費税申告", "消費税の達人\n（ライセンス費用別途）", "freee会計に付帯（追加費用なし）"],
    ["内訳書・概況書", "内訳概況書の達人\n（ライセンス費用別途）", "freee申告 内訳書・概況書\n（別途契約）"],
    ["償却資産申告", "減価償却の達人\n（ライセンス費用別途）", "freee申告 年額59,800円"],
]

for row_idx, row_data in enumerate(data2, 12):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws8.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

ws8['A18'] = "【コスト比較のポイント】"
ws8['A18'].font = Font(bold=True)
ws8['A19'] = "・達人シリーズを既に導入済みの事務所は追加コストなしでMFと連携可能"
ws8['A20'] = "・新規で達人シリーズを導入する場合はライセンス費用を考慮"
ws8['A21'] = "・freeeは申告機能が細分化されており、必要な機能ごとに契約が必要"

ws8.column_dimensions['A'].width = 20
ws8.column_dimensions['B'].width = 28
ws8.column_dimensions['C'].width = 28
ws8.column_dimensions['D'].width = 15
ws8.column_dimensions['E'].width = 15

# ========== シート11: 市場シェア ==========
ws9 = wb.create_sheet("市場シェア")

ws9.merge_cells('A1:C1')
ws9['A1'] = "クラウド会計ソフト市場シェア（2025年）"
ws9['A1'].font = Font(bold=True, size=14)

ws9['A3'] = "【法人向けクラウド会計シェア】"
ws9['A3'].font = font_header

headers = ["ソフト", "シェア", "順位"]
for col, header in enumerate(headers, 1):
    cell = ws9.cell(row=4, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data = [
    ["freee", "32.3%", "1位"],
    ["マネーフォワード", "19.2%", "2位"],
    ["弥生シリーズ", "15.4%", "3位"],
    ["その他", "33.1%", "-"],
]

for row_idx, row_data in enumerate(data, 5):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws9.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws9['A11'] = "【個人事業主向けシェア】"
ws9['A11'].font = font_header

for col, header in enumerate(headers, 1):
    cell = ws9.cell(row=12, column=col, value=header)
    cell.font = font_header
    cell.fill = header_fill_gray
    cell.border = thin_border

data = [
    ["弥生シリーズ", "55.4%", "1位"],
    ["freee", "24.0%", "2位"],
    ["マネーフォワード", "14.3%", "3位"],
    ["その他", "6.3%", "-"],
]

for row_idx, row_data in enumerate(data, 13):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws9.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws9['A19'] = "【会計事務所での導入状況】"
ws9['A19'].font = Font(bold=True)
ws9['A20'] = "・マネーフォワード: TOP500会計事務所の導入数No.1"
ws9['A21'] = "・全国50,000名超の士業とパートナーシップ"
ws9['A22'] = "出典: MM総研 2025年3月末調査"

ws9.column_dimensions['A'].width = 25
ws9.column_dimensions['B'].width = 12
ws9.column_dimensions['C'].width = 10

# ========== シート12: 推奨シナリオ ==========
ws10 = wb.create_sheet("推奨シナリオ")

ws10.merge_cells('A1:B1')
ws10['A1'] = "推奨シナリオ"
ws10['A1'].font = Font(bold=True, size=14)

ws10['A3'] = "マネーフォワード + 達人シリーズ が向いているケース"
ws10['A3'].font = Font(bold=True, size=12, color="3B7DE9")

mf_cases = [
    "達人シリーズを既に導入済み",
    "経理経験者・税理士が操作する",
    "従来の会計ソフトからの移行",
    "銀行・カード連携を重視（2,300以上対応）",
    "中規模以上の法人顧問先",
    "SmartHR等の労務システムと連携したい",
    "内訳書・概況書の自動連携を重視",
    "会計事務所主導で記帳代行",
    "安定した税務申告ワークフローを維持したい",
]

for idx, case in enumerate(mf_cases, 4):
    ws10[f'A{idx}'] = f"・{case}"

ws10['B3'] = "freee + freee申告 が向いているケース"
ws10['B3'].font = Font(bold=True, size=12, color="00AA4F")

freee_cases = [
    "達人シリーズを持っていない（新規導入）",
    "簿記知識のない経営者が自ら入力",
    "スタートアップ・小規模法人",
    "取引数が多く自動化を重視",
    "スマホ完結を求める",
    "POSレジ・EC連携が必要",
    "オールインワンでシンプルに運用",
    "API連携で独自システムと接続",
    "ソフトを一本化してコスト管理したい",
]

for idx, case in enumerate(freee_cases, 4):
    ws10[f'B{idx}'] = f"・{case}"

ws10['A15'] = "【会計事務所としての結論】"
ws10['A15'].font = Font(bold=True, size=12)
ws10.merge_cells('A16:B16')
ws10['A16'] = "達人シリーズを既に導入済みの事務所であれば、MF+達人の組み合わせは申し分ない。"
ws10.merge_cells('A17:B17')
ws10['A17'] = "内訳書・概況書もMFから達人へ科目残高・月別残高を自動連携でき、効率的に作成可能。"
ws10.merge_cells('A18:B18')
ws10['A18'] = "達人を持っていない場合や新規導入の場合は、freee+freee申告の方がコスト面で有利。"
ws10.merge_cells('A19:B19')
ws10['A19'] = "顧問先の規模・業種・ITリテラシーに応じて使い分けるのが現実的。"

ws10.column_dimensions['A'].width = 48
ws10.column_dimensions['B'].width = 48

# 保存
output_path = r"C:\Users\yiwao\business-application\ms-expense-app\docs\mf-freee-comparison.xlsx"
wb.save(output_path)
print(f"Excelファイルを作成しました: {output_path}")
